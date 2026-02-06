# -*- coding: utf-8 -*-
"""
Guardian Life Career Scraper - GitHub Actions Version
Runs automatically on scheduled days via GitHub Actions
"""

from bs4 import BeautifulSoup
import requests
from tqdm import tqdm
import pandas as pd
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import logging

# ============================================================================
# CONFIGURATION - GitHub Actions Version
# ============================================================================

# Schedule is handled by GitHub Actions cron
# See .github/workflows/scraper.yml to change schedule

# Path Configuration (local folders, not Google Drive)
OUTPUT_FOLDER = 'output'
LOG_FOLDER = 'logs'

# Export Configuration
EXPORT_CONFIG = {
    'save_excel': True,
    'save_csv': True,
    'save_json': True,  # Enabled for GitHub history
}

# ============================================================================
# SETUP AND UTILITY FUNCTIONS
# ============================================================================

def setup_folders():
    """Create necessary folders if they don't exist"""
    folders = [OUTPUT_FOLDER, LOG_FOLDER]
    for folder in folders:
        if not os.path.exists(folder):
            os.makedirs(folder)

def setup_logging():
    """Setup logging configuration"""
    log_path = os.path.join(LOG_FOLDER, f'scraper_{get_timestamp()}.log')
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_path),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def get_timestamp():
    """Get current timestamp in IST"""
    return (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d_%H-%M-%S')

def get_date_only():
    """Get current date in IST"""
    return (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime('%Y-%m-%d')

def save_run_history(status, records_count=0, error=None):
    """Save run history to JSON file"""
    history_file = os.path.join(LOG_FOLDER, 'run_history.json')
    
    history_entry = {
        'timestamp': get_timestamp(),
        'date': get_date_only(),
        'status': status,
        'records_scraped': records_count,
        'error': str(error) if error else None
    }
    
    history = []
    if os.path.exists(history_file):
        try:
            with open(history_file, 'r') as f:
                history = json.load(f)
        except:
            history = []
    
    history.append(history_entry)
    
    with open(history_file, 'w') as f:
        json.dump(history, f, indent=2)

# ============================================================================
# SCRAPING FUNCTIONS
# ============================================================================

cookies = {
    'PLAY_SESSION': 'ed1dd99ad8309df39955ca94d4339425751d3c11-guardianlife_pSessionId=iruc1382sugs7n6a53hn84di5b&instance=vps-prod-ie1iirjt.prod-vps.pr503.cust.pdx.wd',
    'wd-browser-id': 'f2fc685c-8c80-4fae-8bea-36a27ee20867',
    'CALYPSO_CSRF_TOKEN': 'd1382620-21d7-4837-8b5b-310c71c84649',
    'wday_vps_cookie': '132946954.53810.0000',
    'timezoneOffset': '-330',
}

headers = {
    'accept': 'application/json',
    'accept-language': 'en-US',
    'content-type': 'application/json',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
}

def role_list_collect(offset):
    """Fetch job listings with offset"""
    json_data = {
        'appliedFacets': {},
        'limit': 20,
        'offset': offset,
        'searchText': '',
    }
    
    try:
        response = requests.post(
            'https://guardianlife.wd5.myworkdayjobs.com/wday/cxs/guardianlife/Guardian-Life-Careers/jobs',
            cookies=cookies,
            headers=headers,
            json=json_data,
            timeout=30
        )
        response.raise_for_status()
        return response.json()
    except Exception as e:
        logger.error(f"Error fetching role list at offset {offset}: {e}")
        return {'jobPostings': []}

def extract_text_from_html(html_string):
    """Extract clean text from HTML"""
    if pd.isna(html_string) or not html_string:
        return ""
    soup = BeautifulSoup(html_string, 'html.parser')
    return soup.get_text(separator=' ', strip=True)

def role_details_fetch(perma):
    """Fetch detailed job information"""
    try:
        response = requests.get(
            f'https://guardianlife.wd5.myworkdayjobs.com/wday/cxs/guardianlife/Guardian-Life-Careers{perma}',
            cookies=cookies,
            headers=headers,
            timeout=30
        )
        response.raise_for_status()
        return response.json()
    except Exception as e:
        logger.error(f"Error fetching details for {perma}: {e}")
        return {}

def scrape_jobs():
    """Main scraping function"""
    logger.info("Starting job scraping...")
    
    collect_list_roles = []
    for i in tqdm(range(0, 500, 20), desc='Collecting Roles'):
        results = role_list_collect(i)
        if 'jobPostings' in results:
            collect_list_roles.extend(results['jobPostings'])
        if not results.get('jobPostings'):
            break
    
    logger.info(f"Collected {len(collect_list_roles)} job listings")
    
    if not collect_list_roles:
        logger.warning("No jobs found")
        return None
    
    roles_df = pd.json_normalize(collect_list_roles).drop_duplicates(['bulletFields'])
    
    collect_role_details = []
    for perma in tqdm(roles_df['externalPath'], desc='Collecting Details'):
        result = role_details_fetch(perma)
        if result:
            result['perma'] = perma
            collect_role_details.append(result)
    
    roledetails_df = pd.json_normalize(collect_role_details)
    
    if 'jobPostingInfo.jobDescription' in roledetails_df.columns:
        roledetails_df['jobPostingInfo.jobDescription'] = roledetails_df['jobPostingInfo.jobDescription'].apply(extract_text_from_html)
    
    final_df = pd.merge(roles_df, roledetails_df, left_on='externalPath', right_on='perma', how='left')
    
    column_mapping = {
        'jobPostingInfo.title': 'Job Title',
        'jobPostingInfo.jobDescription': 'Job Description',
        'jobPostingInfo.location': 'Location',
        'jobPostingInfo.additionalLocations': 'Additional Locations',
        'jobPostingInfo.startDate': 'Start Date',
        'jobPostingInfo.jobReqId': 'Job ID',
        'jobPostingInfo.remoteType': 'Remote Type',
        'jobPostingInfo.externalUrl': 'Application URL'
    }
    
    available_cols = [col for col in column_mapping.keys() if col in final_df.columns]
    final_df = final_df[available_cols]
    final_df = final_df.rename(columns=column_mapping)
    
    final_df.insert(0, 'Scraped Date', get_date_only())
    final_df.insert(1, 'Scraped Time', get_timestamp())
    
    logger.info(f"Successfully processed {len(final_df)} job records")
    
    return final_df

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

def format_excel(file_path):
    """Apply professional formatting to Excel file"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    column_widths = {
        'A': 12, 'B': 18, 'C': 35, 'D': 60, 'E': 20,
        'F': 20, 'G': 12, 'H': 15, 'I': 15, 'J': 50,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30
    
    wb.save(file_path)
    logger.info(f"Excel formatting applied to {file_path}")

def export_data(df):
    """Export data in configured formats"""
    timestamp = get_timestamp()
    date_only = get_date_only()
    
    exported_files = []
    
    if EXPORT_CONFIG['save_excel']:
        excel_filename = f'GuardianLife_Jobs_{date_only}.xlsx'
        excel_path = os.path.join(OUTPUT_FOLDER, excel_filename)
        
        df.to_excel(excel_path, index=False, engine='openpyxl')
        format_excel(excel_path)
        exported_files.append(excel_path)
        logger.info(f"Excel file saved: {excel_filename}")
    
    if EXPORT_CONFIG['save_csv']:
        csv_filename = f'GuardianLife_Jobs_{date_only}.csv'
        csv_path = os.path.join(OUTPUT_FOLDER, csv_filename)
        
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        exported_files.append(csv_path)
        logger.info(f"CSV file saved: {csv_filename}")
    
    if EXPORT_CONFIG['save_json']:
        json_filename = f'GuardianLife_Jobs_{date_only}.json'
        json_path = os.path.join(OUTPUT_FOLDER, json_filename)
        
        df.to_json(json_path, orient='records', indent=2)
        exported_files.append(json_path)
        logger.info(f"JSON file saved: {json_filename}")
    
    return exported_files

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    global logger
    
    setup_folders()
    logger = setup_logging()
    
    logger.info("="*70)
    logger.info("Guardian Life Career Scraper - GitHub Actions Auto Run")
    logger.info("="*70)
    
    try:
        df = scrape_jobs()
        
        if df is None or len(df) == 0:
            logger.warning("No data scraped")
            save_run_history('no_data')
            print("\n⚠️  No jobs found")
            return
        
        exported_files = export_data(df)
        
        logger.info("="*70)
        logger.info(f"✅ Scraping completed successfully!")
        logger.info(f"Total jobs scraped: {len(df)}")
        logger.info(f"Files exported: {len(exported_files)}")
        for file in exported_files:
            logger.info(f"  - {os.path.basename(file)}")
        logger.info("="*70)
        
        save_run_history('success', len(df))
        
        print(f"\n{'='*70}")
        print(f"✅ SUCCESS!")
        print(f"{'='*70}")
        print(f"Jobs scraped: {len(df)}")
        print(f"Files saved in: {OUTPUT_FOLDER}/")
        for file in exported_files:
            print(f"  📄 {os.path.basename(file)}")
        print(f"{'='*70}\n")
        
        return df
        
    except Exception as e:
        logger.error(f"Error during execution: {e}", exc_info=True)
        save_run_history('error', error=e)
        print(f"\n❌ Error: {e}")
        raise

if __name__ == "__main__":
    main()
